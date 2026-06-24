#!/usr/bin/env python3
# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl>=3.1"]
# ///
"""Excel preprocessor: convert xlsx/xls sheets to readable text and emit LLM-ready JSON.

Usage:
  uv run scripts/preprocess_excel.py <source-file>

Outputs a single JSON object to stdout:
  {"text": "...", "metadata": {"format": "xlsx", "source_ref": "<path>", "sheets": N, "rows": N}}
"""

import json
import sys
from pathlib import Path

import openpyxl


def _sheet_to_text(sheet) -> str:
    lines = [f"[Sheet: {sheet.title}]"]
    for row in sheet.iter_rows(values_only=True):
        cells = [str(c) if c is not None else "" for c in row]
        if any(cells):
            lines.append(" | ".join(cells))
    return "\n".join(lines)


def main():
    if len(sys.argv) != 2:
        print("Usage: preprocess_excel.py <source-file>", file=sys.stderr)
        sys.exit(1)

    source_file = Path(sys.argv[1])
    if not source_file.exists():
        print(f"Error: {source_file} not found", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(str(source_file), data_only=True)
    parts: list[str] = []
    total_rows = 0

    for sheet in wb.worksheets:
        parts.append(_sheet_to_text(sheet))
        total_rows += sheet.max_row or 0

    print(json.dumps({
        "text": "\n\n".join(parts),
        "metadata": {
            "format": source_file.suffix.lstrip(".").lower(),
            "source_ref": str(source_file),
            "sheets": len(wb.worksheets),
            "rows": total_rows,
        },
    }))


if __name__ == "__main__":
    main()
